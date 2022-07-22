import numpy as np
from openpyxl import load_workbook

"""Parameters file"""
MAX_SIZE = 4 * 1e6  # 4TB with MB as baseline
REPLICATION_FACTOR = 3
COST_DYNAMO_WRITE_UNIT = 1.4842 / 1e6  # cost per million write requests: 1.4842 / 1e6
COST_DYNAMO_READ_UNIT = 0.2968 / 1e6  # cost per million read requests: 0.2968 / 1e6
COST_DYNAMO_STORAGE = (
    0.29715 / 30 / 24 / 1e3
)  # 0.29715 cost per GB per month -> per Mega byte per hour

# period (hours) of stability over which the optimizer calculates the total cost
WORKLOAD_STABILITY = 10

wb = load_workbook("../selection.xlsx")
ws = wb["Sheet1"]
vm_types = ws["B2:B24"]
vm_types = [cell[0].value for cell in vm_types]
# print(vm_types)
# it is unclear why the cell elements are tuples with the second value = Nil,
# but using [0] works...
# Here I can do this trick because columns are adjacent
vm_bandwidths = ws["B2:C24"]
vm_bandwidths = {vm.value: bandwidth.value for (vm, bandwidth) in vm_bandwidths}
vm_IOPS = ws["D2:D24"]
vm_IOPS = [cell[0].value for cell in vm_IOPS]
vm_IOPS = {vmtype: iops for vmtype, iops in zip(vm_types, vm_IOPS)}
# print(vm_bandwidths)
# print(vm_IOPS)
vm_costs = ws["F2:F24"]
vm_costs = [cell[0].value for cell in vm_costs]
vm_costs = {vmtype: iops for vmtype, iops in zip(vm_types, vm_costs)}
# print(vm_types)
# print()
# print(vm_bandwidths)
# print()
# print(vm_costs)
# print()
# print(vm_IOPS)
# AWS volume pricing
# General Purpose SSD (gp3) - Storage	$0.0924/GB-month
# General Purpose SSD (gp3) - IOPS	3,000 IOPS free and $0.0058/provisioned IOPS-month over 3,000
# General Purpose SSD (gp3) - Throughput	125 MB/s free and $0.0462/provisioned MB/s-month over 125

volumes = ["gp2", "gp3", "io1", "st1", "sc1"]
# volumes = ["gp2"]

COST_VOLUME_STORAGE = [
    0.1155,
    0.0924,
    0.1449,
    0.525,
    0.01764,
]  # / 30 / 24 / 1e3  # per MB per hour
COST_VOLUME_STORAGE = np.multiply(COST_VOLUME_STORAGE, 1 / 30 / 24 / 1e3)
COST_VOLUME_STORAGE = dict(zip(volumes, COST_VOLUME_STORAGE))

COST_VOLUME_IOPS = [0, 0.0058, 0.0756, 0, 0]  # per IOPS-month
COST_VOLUME_IOPS = np.multiply(COST_VOLUME_IOPS, 1 / 30 / 24)  # per IOPS-hour
COST_VOLUME_IOPS = dict(zip(volumes, COST_VOLUME_IOPS))

COST_VOLUME_THROUGHPUT = [0, 0.0462 / 30 / 24, 0, 0, 0]  # $ per MB/s per hour
COST_VOLUME_THROUGHPUT = dict(zip(volumes, COST_VOLUME_THROUGHPUT))

# MB to MiB --> * 10^6 / 2^20
# MiB to MB --> * 2^20 / 10^6
# General Purpose SSD (gp3) - Storage	    $0.0924/GB-month
# General Purpose SSD (gp3) - IOPS	        3,000 IOPS free and $0.0058/provisioned IOPS-month over 3,000
# General Purpose SSD (gp3) - Throughput	125 MB/s free and $0.0462/provisioned MB/s-month over 125
# General Purpose SSD (gp2) Volumes	        $0.1155 per GB-month of provisioned storage
# Provisioned IOPS SSD (io1) Volumes	    $0.1449 per GB-month of provisioned storage AND $0.0756 per provisioned IOPS-month
# Throughput Optimized HDD (st1) Volumes	$0.0525 per GB-month of provisioned storage
# Cold HDD (sc1) Volumes                    $0.01764 per GB-month of provisioned storage
MAX_VOLUME_IOPS = [
    16000,
    16000,
    64000,
    500,
    250,
]  # 16 KiB I/O for SSD(gp2,gp3,io1), 1MiB I/O for HDD(sc1,st1)
IO_FACTOR = [
    10**6 / 2**10 / 16,  # MB to KiB, then 1 I/O every 16
    10**6 / 2**10 / 16,  # MB to KiB, then 1 I/O every 16
    10**6 / 2**10 / 16,  # MB to KiB, then 1 I/O every 16
    10**6 / 2**20,  # MB to MiB, then 1 I/O every MB
    10**6 / 2**20,  # MB to MiB, then 1 I/O every MB
]
IO_FACTOR = dict(zip(volumes, IO_FACTOR))
MAX_VOLUME_IOPS = dict(zip(volumes, MAX_VOLUME_IOPS))
MAX_VOLUME_THROUGHPUT = [250, 1000, 1000, 500, 250]  # MiB/s
MAX_VOLUME_THROUGHPUT = dict(zip(volumes, MAX_VOLUME_THROUGHPUT))
